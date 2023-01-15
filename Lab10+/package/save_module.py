# -*- coding: utf-8 -*-

from openpyxl import *


def data_Save(*args):
    # wb = openpyxl.load_workbook("Result.xlsx", read_only=False)
    wb = load_workbook("Result.xlsx")
    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells

    ws['A1'] = "Обои:"
    ws['B1'] = "|"
    ws['C1'] = "Плитка:"
    ws['D1'] = "|"
    ws['E1'] = "Ламинат:"
    ws['F1'] = "|"

    if args[0] != 0:
        ws['A2'] = "Количество:"
        ws['B2'] = args[0]

        ws['A3'] = "Общая цена:"
        ws['B3'].value = args[1]
        wb.save("Result.xlsx")
    elif args[2] != 0:
        ws['C2'] = "Количество:"
        ws['D2'] = args[2]

        ws['C3'] = "Общая цена:"
        ws['D3'].value = args[3]
        wb.save("Result.xlsx")
    else:
        ws['E2'] = "Количество:"
        ws['F2'] = args[4]

        ws['E3'] = "Общая цена:"
        ws['F3'].value = args[5]
        wb.save("Result.xlsx")
